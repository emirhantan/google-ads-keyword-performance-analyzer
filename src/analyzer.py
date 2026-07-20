# -*- coding: utf-8 -*-
"""
Google Ads Keyword Performance Analyzer

A configurable decision-support tool that:
- validates Google Ads keyword exports,
- separates low-data keywords using a configurable spend threshold,
- calculates continuous 0-100 KPI scores,
- applies AHP weights,
- runs sensitivity analysis,
- creates a management-ready Excel dashboard.

No company-specific data is included in this repository.
"""

from __future__ import annotations

import argparse
from bisect import bisect_left, bisect_right
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REQUIRED_COLUMNS = [
    "Arama anahtar kelimesi",
    "Göstr.",
    "Tıklamalar",
    "TO",
    "Etkileşim sağlanan oturum yüzdesi (GA4)",
    "Maliyet",
    "Ort. TBM",
    "Oturum başına ort. etkileşim süresi (saniye) (GA4)",
    "Dönüşümler",
]

CRITERIA = ["CTR", "GA4", "TBM", "Sure"]

AHP_MATRIX = np.array(
    [
        [1,   1 / 5, 3,   1 / 5],
        [5,   1,     3,   1 / 3],
        [1 / 3, 1 / 3, 1, 1 / 9],
        [5,   3,     9,   1],
    ],
    dtype=float,
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Analyze Google Ads keyword performance and build an Excel dashboard."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("sample_data/google_ads_sample.xlsx"),
        help="Path to the Google Ads keyword export.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("sample_output/sample_dashboard.xlsx"),
        help="Path for the generated Excel report.",
    )
    parser.add_argument(
        "--min-spend",
        type=float,
        default=500.0,
        help="Minimum spend required for the main ranking.",
    )
    return parser.parse_args()


def safe_numeric(value):
    if pd.isna(value):
        return np.nan
    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value)

    text = (
        str(value)
        .strip()
        .replace("%", "")
        .replace("₺", "")
        .replace("TL", "")
        .replace("\u00a0", "")
        .replace(" ", "")
    )

    if not text:
        return np.nan

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")

    return pd.to_numeric(text, errors="coerce")


def percentile_score(series, reference, higher_is_better=True):
    reference_values = reference.dropna().sort_values().to_numpy()
    n_ref = len(reference_values)

    if n_ref == 0:
        return pd.Series(np.nan, index=series.index)

    scores = []
    for value in series:
        if pd.isna(value):
            scores.append(np.nan)
            continue

        left = bisect_left(reference_values, value)
        right = bisect_right(reference_values, value)
        midpoint_rank = (left + right) / 2
        score = 100 * midpoint_rank / n_ref

        if not higher_is_better:
            score = 100 - score

        scores.append(max(0.0, min(100.0, score)))

    return pd.Series(scores, index=series.index)


def calculate_ahp():
    eigenvalues, eigenvectors = np.linalg.eig(AHP_MATRIX)
    main_index = int(np.argmax(eigenvalues.real))
    lambda_max = float(eigenvalues[main_index].real)

    weights = eigenvectors[:, main_index].real
    weights = weights / weights.sum()

    weight_series = pd.Series(weights, index=CRITERIA, name="AHP_Agirligi")

    n = len(CRITERIA)
    ci = (lambda_max - n) / (n - 1)
    ri = 0.90
    cr = ci / ri

    if cr >= 0.10:
        raise ValueError(f"AHP pairwise comparisons are inconsistent. CR={cr:.4f}")

    return weight_series, lambda_max, ci, cr


def load_and_validate(input_file: Path) -> pd.DataFrame:
    if not input_file.exists():
        raise FileNotFoundError(f"Input file was not found: {input_file}")

    df = pd.read_excel(input_file)
    missing = [column for column in REQUIRED_COLUMNS if column not in df.columns]

    if missing:
        raise KeyError("Missing required columns:\n- " + "\n- ".join(missing))

    keyword = df[REQUIRED_COLUMNS].copy()

    numeric_columns = [column for column in REQUIRED_COLUMNS if column != "Arama anahtar kelimesi"]
    for column in numeric_columns:
        keyword[column] = keyword[column].map(safe_numeric)

    keyword = keyword.dropna(subset=["Arama anahtar kelimesi"]).copy()
    keyword["Arama anahtar kelimesi"] = (
        keyword["Arama anahtar kelimesi"].astype(str).str.strip()
    )

    for column in ["TO", "Etkileşim sağlanan oturum yüzdesi (GA4)"]:
        non_null = keyword[column].dropna()
        if not non_null.empty and non_null.max() > 1.5:
            keyword[column] = keyword[column] / 100

    kpi_columns = [
        "TO",
        "Etkileşim sağlanan oturum yüzdesi (GA4)",
        "Ort. TBM",
        "Oturum başına ort. etkileşim süresi (saniye) (GA4)",
    ]
    keyword = keyword.dropna(subset=kpi_columns).copy()

    calculated_ctr = keyword["Tıklamalar"] / keyword["Göstr."].replace(0, np.nan)
    calculated_cpc = keyword["Maliyet"] / keyword["Tıklamalar"].replace(0, np.nan)

    max_ctr_difference = (keyword["TO"] - calculated_ctr).abs().max(skipna=True)
    max_cpc_difference = (keyword["Ort. TBM"] - calculated_cpc).abs().max(skipna=True)

    print(f"Maximum CTR validation difference: {max_ctr_difference:.6f}")
    print(f"Maximum CPC validation difference: {max_cpc_difference:.6f}")

    return keyword


def analyze(keyword: pd.DataFrame, min_spend: float):
    keyword = keyword.copy()
    keyword["Veri_Yeterliligi"] = np.where(
        keyword["Maliyet"] >= min_spend,
        "Yeterli Veri",
        "İzlemede / Yetersiz Veri",
    )

    sufficient_mask = keyword["Maliyet"] >= min_spend
    sufficient = keyword.loc[sufficient_mask].copy()

    if len(sufficient) < 10:
        raise ValueError(
            "Too few keywords meet the data-sufficiency threshold. "
            "Lower --min-spend or provide a larger dataset."
        )

    weights, lambda_max, ci, cr = calculate_ahp()

    keyword["CTR_Skoru"] = percentile_score(
        keyword["TO"], sufficient["TO"], higher_is_better=True
    )
    keyword["GA4_Skoru"] = percentile_score(
        keyword["Etkileşim sağlanan oturum yüzdesi (GA4)"],
        sufficient["Etkileşim sağlanan oturum yüzdesi (GA4)"],
        higher_is_better=True,
    )
    keyword["TBM_Skoru"] = percentile_score(
        keyword["Ort. TBM"], sufficient["Ort. TBM"], higher_is_better=False
    )
    keyword["Sure_Skoru"] = percentile_score(
        keyword["Oturum başına ort. etkileşim süresi (saniye) (GA4)"],
        sufficient["Oturum başına ort. etkileşim süresi (saniye) (GA4)"],
        higher_is_better=True,
    )

    keyword["Performans_Endeksi"] = (
        keyword["CTR_Skoru"] * weights["CTR"]
        + keyword["GA4_Skoru"] * weights["GA4"]
        + keyword["TBM_Skoru"] * weights["TBM"]
        + keyword["Sure_Skoru"] * weights["Sure"]
    ).round(2)

    sufficient_index = keyword.loc[sufficient_mask, "Performans_Endeksi"]
    q1 = float(sufficient_index.quantile(0.25))
    median = float(sufficient_index.quantile(0.50))
    q3 = float(sufficient_index.quantile(0.75))

    def performance_level(row):
        if row["Maliyet"] < min_spend:
            return "İzlemede / Yetersiz Veri"
        if row["Performans_Endeksi"] >= q3:
            return "Lider"
        if row["Performans_Endeksi"] >= median:
            return "Güçlü"
        if row["Performans_Endeksi"] >= q1:
            return "Geliştirilebilir"
        return "Öncelikli İnceleme"

    keyword["Performans_Seviyesi"] = keyword.apply(performance_level, axis=1)

    score_mapping = {
        "CTR": "CTR_Skoru",
        "GA4": "GA4_Skoru",
        "TBM": "TBM_Skoru",
        "Süre": "Sure_Skoru",
    }

    keyword["Guclu_Yon"] = keyword.apply(
        lambda row: ", ".join(
            label for label, column in score_mapping.items() if row[column] >= 75
        ) or "-",
        axis=1,
    )
    keyword["Gelisim_Alani"] = keyword.apply(
        lambda row: ", ".join(
            label for label, column in score_mapping.items() if row[column] < 50
        ) or "-",
        axis=1,
    )

    def priority(row):
        if row["Maliyet"] < min_spend:
            return "İzlemede"
        if row["Performans_Seviyesi"] == "Öncelikli İnceleme":
            return "Yüksek"
        if row["Performans_Seviyesi"] == "Geliştirilebilir":
            return "Orta"
        return "Düşük"

    keyword["Oncelik"] = keyword.apply(priority, axis=1)
    keyword["Performans_Sirasi"] = np.nan

    sufficient_ranking = keyword.loc[sufficient_mask].sort_values(
        ["Performans_Endeksi", "Maliyet"], ascending=[False, False]
    )
    keyword.loc[
        sufficient_ranking.index, "Performans_Sirasi"
    ] = np.arange(1, len(sufficient_ranking) + 1)

    sensitivity_rows = []
    base_rank = keyword.loc[sufficient_mask, "Performans_Endeksi"].rank(
        ascending=False, method="min"
    )
    base_top10 = set(
        keyword.loc[sufficient_mask]
        .nlargest(10, "Performans_Endeksi")["Arama anahtar kelimesi"]
    )

    for criterion in CRITERIA:
        scenario_weights = weights.copy()
        scenario_weights[criterion] *= 1.10
        scenario_weights = scenario_weights / scenario_weights.sum()

        scenario_score = (
            keyword["CTR_Skoru"] * scenario_weights["CTR"]
            + keyword["GA4_Skoru"] * scenario_weights["GA4"]
            + keyword["TBM_Skoru"] * scenario_weights["TBM"]
            + keyword["Sure_Skoru"] * scenario_weights["Sure"]
        )

        scenario_rank = scenario_score.loc[sufficient_mask].rank(
            ascending=False, method="min"
        )
        rank_difference = (base_rank - scenario_rank).abs()

        scenario_top10 = set(
            keyword.loc[sufficient_mask]
            .assign(Scenario_Score=scenario_score.loc[sufficient_mask])
            .nlargest(10, "Scenario_Score")["Arama anahtar kelimesi"]
        )

        sensitivity_rows.append(
            {
                "Senaryo": f"{criterion} ağırlığı +%10",
                "Spearman": round(
                    float(
                        pd.concat([base_rank, scenario_rank], axis=1)
                        .corr(method="spearman")
                        .iloc[0, 1]
                    ),
                    4,
                ),
                "Ortalama Sıra Değişimi": round(float(rank_difference.mean()), 2),
                "Maksimum Sıra Değişimi": int(rank_difference.max()),
                "İlk 10 Ortak": f"{len(base_top10 & scenario_top10)}/10",
            }
        )

    methodology = {
        "weights": weights,
        "lambda_max": lambda_max,
        "ci": ci,
        "cr": cr,
        "q1": q1,
        "median": median,
        "q3": q3,
        "sensitivity": pd.DataFrame(sensitivity_rows),
        "correlation": keyword[
            [
                "TO",
                "Etkileşim sağlanan oturum yüzdesi (GA4)",
                "Ort. TBM",
                "Oturum başına ort. etkileşim süresi (saniye) (GA4)",
            ]
        ].corr(),
    }

    return keyword, methodology


def build_report(keyword, methodology, output_file: Path, min_spend: float):
    output_file.parent.mkdir(parents=True, exist_ok=True)

    result = keyword[
        [
            "Arama anahtar kelimesi",
            "Göstr.",
            "Tıklamalar",
            "TO",
            "Etkileşim sağlanan oturum yüzdesi (GA4)",
            "Maliyet",
            "Ort. TBM",
            "Oturum başına ort. etkileşim süresi (saniye) (GA4)",
            "Dönüşümler",
            "Veri_Yeterliligi",
            "CTR_Skoru",
            "GA4_Skoru",
            "TBM_Skoru",
            "Sure_Skoru",
            "Performans_Endeksi",
            "Performans_Sirasi",
            "Performans_Seviyesi",
            "Guclu_Yon",
            "Gelisim_Alani",
            "Oncelik",
        ]
    ].copy()

    result.columns = [
        "Arama anahtar kelimesi",
        "Gösterim",
        "Tıklamalar",
        "CTR",
        "GA4 Etkileşim",
        "Toplam Maliyet",
        "Ort. TBM",
        "Ort. Etkileşim Süresi",
        "Dönüşümler",
        "Veri Yeterliliği",
        "CTR Skoru",
        "GA4 Skoru",
        "TBM Skoru",
        "Süre Skoru",
        "Performans Endeksi",
        "Performans Sırası",
        "Performans Seviyesi",
        "Güçlü Yön",
        "Gelişim Alanı",
        "Öncelik",
    ]

    result["Ranking_Group"] = np.where(result["Toplam Maliyet"] >= min_spend, 1, 0)
    result = result.sort_values(
        ["Ranking_Group", "Performans Endeksi", "Toplam Maliyet"],
        ascending=[False, False, False],
    ).drop(columns=["Ranking_Group"]).reset_index(drop=True)

    sufficient = result[result["Toplam Maliyet"] >= min_spend].copy()
    monitoring = result[result["Toplam Maliyet"] < min_spend].copy()

    top10 = sufficient.nlargest(10, "Performans Endeksi")[
        [
            "Arama anahtar kelimesi",
            "Performans Endeksi",
            "Performans Seviyesi",
            "Toplam Maliyet",
        ]
    ]
    bottom10 = sufficient.nsmallest(10, "Performans Endeksi")[
        [
            "Arama anahtar kelimesi",
            "Performans Endeksi",
            "Gelişim Alanı",
            "Toplam Maliyet",
        ]
    ]
    expensive10 = result.nlargest(10, "Toplam Maliyet")[
        [
            "Arama anahtar kelimesi",
            "Toplam Maliyet",
            "Ort. TBM",
            "Performans Seviyesi",
        ]
    ]
    monitoring10 = monitoring.nlargest(10, "Performans Endeksi")[
        [
            "Arama anahtar kelimesi",
            "Performans Endeksi",
            "Toplam Maliyet",
            "Güçlü Yön",
        ]
    ]

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        result.to_excel(writer, sheet_name="Anahtar Kelime Analizi", index=False)
        methodology["correlation"].to_excel(
            writer, sheet_name="Metodoloji", startrow=1, startcol=0
        )
        methodology["weights"].rename("AHP Ağırlığı").to_frame().to_excel(
            writer, sheet_name="Metodoloji", startrow=8, startcol=0
        )
        methodology["sensitivity"].to_excel(
            writer, sheet_name="Metodoloji", startrow=15, startcol=0, index=False
        )
        pd.DataFrame().to_excel(writer, sheet_name="Dashboard", index=False)

    wb = load_workbook(output_file)
    ws_analysis = wb["Anahtar Kelime Analizi"]
    ws_method = wb["Metodoloji"]
    ws_dashboard = wb["Dashboard"]

    colors = {
        "navy": "17365D",
        "blue": "1F4E78",
        "light_blue": "D9EAF7",
        "green": "C6EFCE",
        "light_green": "E2F0D9",
        "yellow": "FFF2CC",
        "red": "FFC7CE",
        "gray": "E7E6E6",
        "light_gray": "F5F7FA",
        "white": "FFFFFF",
    }

    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor=colors["blue"])
    header_font = Font(color=colors["white"], bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws_analysis.freeze_panes = "B2"
    ws_analysis.auto_filter.ref = ws_analysis.dimensions
    ws_analysis.sheet_view.showGridLines = False

    header_map = {cell.value: cell.column for cell in ws_analysis[1]}

    for cell in ws_analysis[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for row_number in range(2, ws_analysis.max_row + 1):
        row_fill = colors["light_gray"] if row_number % 2 == 0 else colors["white"]
        for column_number in range(1, ws_analysis.max_column + 1):
            cell = ws_analysis.cell(row_number, column_number)
            cell.fill = PatternFill("solid", fgColor=row_fill)
            cell.border = border
            cell.alignment = left_wrap

    for heading in ["CTR", "GA4 Etkileşim"]:
        column = header_map[heading]
        for row_number in range(2, ws_analysis.max_row + 1):
            ws_analysis.cell(row_number, column).number_format = "0.00%"

    for heading in ["Toplam Maliyet", "Ort. TBM"]:
        column = header_map[heading]
        for row_number in range(2, ws_analysis.max_row + 1):
            ws_analysis.cell(row_number, column).number_format = '#,##0.00 "TL"'

    time_column = header_map["Ort. Etkileşim Süresi"]
    for row_number in range(2, ws_analysis.max_row + 1):
        ws_analysis.cell(row_number, time_column).number_format = '0.00 "sn"'

    for heading in [
        "CTR Skoru",
        "GA4 Skoru",
        "TBM Skoru",
        "Süre Skoru",
        "Performans Endeksi",
    ]:
        column = header_map[heading]
        letter = get_column_letter(column)
        ws_analysis.conditional_formatting.add(
            f"{letter}2:{letter}{ws_analysis.max_row}",
            ColorScaleRule(
                start_type="min",
                start_color=colors["red"],
                mid_type="percentile",
                mid_value=50,
                mid_color=colors["yellow"],
                end_type="max",
                end_color=colors["green"],
            ),
        )

    ws_analysis.conditional_formatting.add(
        f"A2:T{ws_analysis.max_row}",
        FormulaRule(
            formula=['$J2="İzlemede / Yetersiz Veri"'],
            fill=PatternFill("solid", fgColor=colors["gray"]),
        ),
    )

    level_colors = {
        "Lider": colors["green"],
        "Güçlü": colors["light_green"],
        "Geliştirilebilir": colors["yellow"],
        "Öncelikli İnceleme": colors["red"],
        "İzlemede / Yetersiz Veri": colors["gray"],
    }

    level_column = header_map["Performans Seviyesi"]
    for row_number in range(2, ws_analysis.max_row + 1):
        cell = ws_analysis.cell(row_number, level_column)
        if cell.value in level_colors:
            cell.fill = PatternFill("solid", fgColor=level_colors[cell.value])
            cell.font = Font(bold=True)
            cell.alignment = center

    widths = {
        "Arama anahtar kelimesi": 32,
        "Gösterim": 12,
        "Tıklamalar": 12,
        "CTR": 11,
        "GA4 Etkileşim": 16,
        "Toplam Maliyet": 18,
        "Ort. TBM": 14,
        "Ort. Etkileşim Süresi": 21,
        "Dönüşümler": 12,
        "Veri Yeterliliği": 24,
        "CTR Skoru": 12,
        "GA4 Skoru": 12,
        "TBM Skoru": 12,
        "Süre Skoru": 12,
        "Performans Endeksi": 18,
        "Performans Sırası": 17,
        "Performans Seviyesi": 23,
        "Güçlü Yön": 24,
        "Gelişim Alanı": 24,
        "Öncelik": 12,
    }
    for heading, width in widths.items():
        ws_analysis.column_dimensions[get_column_letter(header_map[heading])].width = width

    # Dashboard
    ws_dashboard.sheet_view.showGridLines = False
    ws_dashboard.merge_cells("A1:L2")
    ws_dashboard["A1"] = "GOOGLE ADS KEYWORD PERFORMANCE DASHBOARD"
    ws_dashboard["A1"].font = Font(
        bold=True, size=20, color=colors["navy"]
    )
    ws_dashboard["A1"].fill = PatternFill("solid", fgColor=colors["light_blue"])
    ws_dashboard["A1"].alignment = center

    cards = [
        ("Total Keywords", len(result)),
        ("Sufficient Data", len(sufficient)),
        ("Monitoring", len(monitoring)),
        ("Leaders", int((sufficient["Performans Seviyesi"] == "Lider").sum())),
        ("Total Spend", float(result["Toplam Maliyet"].sum())),
        ("Conversions", float(result["Dönüşümler"].sum())),
    ]

    for i, (label, value) in enumerate(cards):
        column = 1 + i * 2
        ws_dashboard.merge_cells(
            start_row=4, start_column=column, end_row=4, end_column=column + 1
        )
        ws_dashboard.merge_cells(
            start_row=5, start_column=column, end_row=6, end_column=column + 1
        )
        ws_dashboard.cell(4, column, label)
        ws_dashboard.cell(5, column, value)
        ws_dashboard.cell(4, column).fill = header_fill
        ws_dashboard.cell(4, column).font = header_font
        ws_dashboard.cell(4, column).alignment = center
        ws_dashboard.cell(5, column).fill = PatternFill(
            "solid", fgColor=colors["light_blue"]
        )
        ws_dashboard.cell(5, column).font = Font(
            bold=True, size=15, color=colors["navy"]
        )
        ws_dashboard.cell(5, column).alignment = center

    def write_table(frame, start_row, start_column, title):
        ws_dashboard.cell(start_row, start_column, title)
        ws_dashboard.cell(start_row, start_column).font = Font(
            bold=True, size=13, color=colors["navy"]
        )
        header_row = start_row + 1
        for offset, column_name in enumerate(frame.columns):
            cell = ws_dashboard.cell(header_row, start_column + offset, column_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        for row_offset, (_, data_row) in enumerate(frame.iterrows(), start=1):
            for column_offset, value in enumerate(data_row):
                cell = ws_dashboard.cell(
                    header_row + row_offset, start_column + column_offset, value
                )
                cell.border = border
                cell.alignment = left_wrap

    write_table(top10, 9, 1, "Top 10 - Sufficient Data")
    write_table(bottom10, 9, 6, "Bottom 10 - Sufficient Data")
    write_table(expensive10, 9, 11, "Top 10 by Spend")
    write_table(monitoring10, 23, 1, "High-Potential Monitoring Keywords")

    distribution = (
        result["Performans Seviyesi"]
        .value_counts()
        .rename_axis("Performance Level")
        .reset_index(name="Keyword Count")
    )
    write_table(distribution, 23, 6, "Performance Distribution")

    chart = BarChart()
    chart.type = "bar"
    chart.title = "Top 10 Keywords"
    chart.legend = None
    chart.add_data(
        Reference(ws_dashboard, min_col=2, min_row=10, max_row=20),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(ws_dashboard, min_col=1, min_row=11, max_row=20)
    )
    chart.height = 8
    chart.width = 18
    ws_dashboard.add_chart(chart, "A38")

    pie = PieChart()
    pie.title = "Performance Distribution"
    pie.add_data(
        Reference(
            ws_dashboard,
            min_col=7,
            min_row=24,
            max_row=24 + len(distribution),
        ),
        titles_from_data=True,
    )
    pie.set_categories(
        Reference(
            ws_dashboard,
            min_col=6,
            min_row=25,
            max_row=24 + len(distribution),
        )
    )
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.height = 8
    pie.width = 12
    ws_dashboard.add_chart(pie, "J38")

    # Methodology
    ws_method["A1"] = "KPI Correlation Matrix"
    ws_method["A8"] = "AHP Weights"
    ws_method["D8"] = "AHP Consistency"
    ws_method["D9"] = "Lambda Max"
    ws_method["E9"] = methodology["lambda_max"]
    ws_method["D10"] = "CI"
    ws_method["E10"] = methodology["ci"]
    ws_method["D11"] = "CR"
    ws_method["E11"] = methodology["cr"]
    ws_method["D12"] = "Status"
    ws_method["E12"] = "Consistent"

    for sheet in [ws_dashboard, ws_method]:
        for column in range(1, 18):
            sheet.column_dimensions[get_column_letter(column)].width = 16

    wb.save(output_file)


def main():
    args = parse_args()
    keyword = load_and_validate(args.input)
    analyzed, methodology = analyze(keyword, args.min_spend)
    build_report(analyzed, methodology, args.output, args.min_spend)
    print(f"Report created: {args.output.resolve()}")


if __name__ == "__main__":
    main()
